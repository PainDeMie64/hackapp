import csv
import io

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse

from app.store import get_session
from app.models import Prospect

router = APIRouter(prefix="/api/search")

HEADERS = [
    "Entreprise", "Secteur", "Localisation", "Site web", "Description",
    "Chiffre d'affaires", "Effectifs", "CA/Employé", "Potentiel de croissance",
    "Prestations intellectuelles", "Actualités", "Confiance", "Sources",
]


def prospect_to_row(p: Prospect) -> list[str]:
    return [
        p.name, p.sector, p.location, p.website or "", p.description or "",
        p.revenue or "", p.headcount or "", p.revenue_per_fte or "",
        p.growth_potential or "", p.uses_intellectual_services or "",
        p.news_summary or "", p.confidence, " | ".join(p.source_urls),
    ]


@router.get("/{search_id}/export/csv")
async def export_csv(search_id: str):
    entry = get_session(search_id)
    if not entry:
        raise HTTPException(status_code=404, detail="Search not found")
    session = entry["session"]

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(HEADERS)
    for p in session.prospects:
        writer.writerow(prospect_to_row(p))

    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=prospects-{search_id[:8]}.csv"},
    )


@router.get("/{search_id}/export/xlsx")
async def export_xlsx(search_id: str):
    entry = get_session(search_id)
    if not entry:
        raise HTTPException(status_code=404, detail="Search not found")
    session = entry["session"]

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Prospects"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0052A5", end_color="0052A5", fill_type="solid")

    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, p in enumerate(session.prospects, 2):
        for col_idx, val in enumerate(prospect_to_row(p), 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=prospects-{search_id[:8]}.xlsx"},
    )
